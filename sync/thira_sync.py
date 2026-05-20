# -*- coding: utf-8 -*-
"""
THIRA Dashboard — Auto Sync
============================
ดึง ใบยกตู้V4TEST.xlsm จาก Dropbox (หรือ local path) → ลง Supabase ants.containers

ENV (สำหรับ GitHub Actions):
  DROPBOX_APP_KEY, DROPBOX_APP_SECRET, DROPBOX_REFRESH_TOKEN
  SUPABASE_URL, SUPABASE_KEY

ENV (สำหรับรันในเครื่อง):
  ใช้ --local <path> แทน Dropbox API
  SUPABASE_URL, SUPABASE_KEY
"""

import os, sys, io, argparse
from datetime import datetime, date

# บังคับ stdout ให้เป็น UTF-8 (รองรับอีโมจิ + ภาษาไทย ทั้ง import และ direct run)
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass


def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def to_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def to_int(v):
    n = to_num(v)
    return int(n) if n is not None else None


def to_str(v):
    if v is None or v == "":
        return None
    return str(v).strip()


def load_workbook_from_dropbox(dbx_path: str):
    import dropbox, openpyxl
    DBX_APP_KEY = os.environ["DROPBOX_APP_KEY"]
    DBX_APP_SEC = os.environ["DROPBOX_APP_SECRET"]
    DBX_REFRESH = os.environ["DROPBOX_REFRESH_TOKEN"]
    dbx = dropbox.Dropbox(
        oauth2_refresh_token=DBX_REFRESH,
        app_key=DBX_APP_KEY,
        app_secret=DBX_APP_SEC,
    )
    log(f"📥 ดาวน์โหลด {dbx_path} จาก Dropbox...")
    _, res = dbx.files_download(dbx_path)
    buf = io.BytesIO(res.content)
    return openpyxl.load_workbook(buf, read_only=True, keep_vba=True, data_only=True)


def load_workbook_local(path: str):
    import openpyxl
    log(f"📁 อ่านไฟล์ local: {path}")
    return openpyxl.load_workbook(path, read_only=True, keep_vba=True, data_only=True)


def parse_database(wb):
    """Parse DATABASE sheet → list of dict ของ ants.containers"""
    if "DATABASE" not in wb.sheetnames:
        log("⚠️  ไม่พบ sheet 'DATABASE'")
        return []
    ws = wb["DATABASE"]
    rows = []
    for i, r in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        # ต้องการ 11 cols: ทะเบียนรถ, จำนวนมัด, โรงงาน(inv), เข้าโรงงาน ว/ท, โรงงาน, สายเรือ, เบอร์ตู้, น.น.ตู้, CADO NO., DATE PICK UP, STATUS
        plate, bundles, inv, date_in, factory, vessel, cont_no, weight, cado, pickup, status = r[:11]
        if all(v in (None, "", 0) for v in (plate, inv, date_in, cont_no)):
            continue
        rows.append({
            "truck_plate": to_str(plate),
            "bundles": to_int(bundles),
            "invoice_no": to_str(inv),
            "date_in": to_date(date_in),
            "factory": to_str(factory),
            "vessel": to_str(vessel),
            "container_no": to_str(cont_no),
            "weight": to_num(weight),
            "cado_no": to_str(cado),
            "date_pickup": to_date(pickup),
            "status": to_str(status),
            "source_row": i,
        })
    return rows


def sync_to_supabase(rows):
    """
    Atomic sync ผ่าน RPC function ants.sync_containers(jsonb)
    ─────────────────────────────────────────────────────────
    function นี้ทำ DELETE+INSERT ใน transaction เดียว + ถือ advisory lock
    → กัน race condition เมื่อ local watcher + GitHub cron รันทับกัน

    Fallback: ถ้า RPC ยังไม่ deploy (function ไม่มี) → ใช้แบบเดิม + warning
    """
    from supabase import create_client
    URL = os.environ["SUPABASE_URL"]
    KEY = os.environ["SUPABASE_KEY"]
    supa = create_client(URL, KEY)

    log(f"📊 มี {len(rows)} rows จะ sync ลง ants.containers")

    # ใช้ RPC atomic function (default)
    try:
        log("  ⚙️  เรียก ants.sync_containers (atomic) ...")
        resp = supa.schema("ants").rpc("sync_containers", {"rows_json": rows}).execute()
        result = resp.data[0] if resp.data else {}
        deleted = result.get("deleted_count", "?")
        inserted = result.get("inserted_count", 0)
        log(f"     ลบ {deleted} rows → insert {inserted} rows  (atomic ✓)")
        return inserted
    except Exception as e:
        msg = str(e)
        # ถ้า function ไม่ exist — fallback ไปแบบเดิม (race-prone)
        if "function" in msg.lower() and ("not exist" in msg.lower() or "not found" in msg.lower() or "PGRST202" in msg):
            log(f"  ⚠️  RPC function ยังไม่ deploy — fallback แบบเดิม (เสี่ยง race)")
            log(f"     → รัน sync_atomic.sql ที่ Supabase Dashboard ก่อน")
        else:
            raise

    # ===== Fallback (เดิม) — ใช้เมื่อยังไม่ได้ deploy RPC =====
    log("  🗑️  ลบ rows เก่าทั้งหมด...")
    del_res = supa.schema("ants").table("containers").delete().gt("id", 0).execute()
    log(f"     ลบไป {len(del_res.data) if del_res.data else '?'} rows")

    inserted = 0
    chunk = 500
    for i in range(0, len(rows), chunk):
        batch = rows[i:i + chunk]
        resp = supa.schema("ants").table("containers").insert(batch).execute()
        inserted += len(resp.data) if resp.data else 0
        log(f"  ✓ insert {inserted}/{len(rows)}")
    return inserted


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--local", help="path to local .xlsm file (skip Dropbox API)")
    args = ap.parse_args()

    if args.local:
        wb = load_workbook_local(args.local)
    else:
        wb = load_workbook_from_dropbox("/ใบยกตู้/ใบยกตู้V4TEST.xlsm")

    rows = parse_database(wb)
    if not rows:
        log("⚠️  ไม่มี rows — skip sync")
        return 0

    n = sync_to_supabase(rows)
    log(f"✅ sync เสร็จ: {n} containers")
    return 0


if __name__ == "__main__":
    sys.exit(main())
