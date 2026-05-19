"""
Import data from Excel files → Supabase
Sources:
  - TEST บัญชีรวมหัวลาก 2026.xlsm  →  thira.{outcomes, incomes, prices, prices_update, fuel_rates}
  - ใบยกตู้V4TEST.xlsm              →  ants.containers

Usage:
  set SUPABASE_URL=https://xxxx.supabase.co
  set SUPABASE_SERVICE_KEY=<service_role key>
  python import_excel.py [--dry-run] [--tables outcomes,incomes,prices,...]
"""
import os
import sys
import io
import argparse
from datetime import datetime, date

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from supabase import create_client, Client

THIRA_XLSM = r'C:\Users\Admin\Dropbox\THITA-RT\TEST บัญชีรวมหัวลาก 2026.xlsm'
ANTS_XLSM  = r'C:\Users\Admin\Dropbox\ใบยกตู้\ใบยกตู้V4TEST.xlsm'

TRUCKS = ['71-9538','72-1620','72-2237','72-2420','72-2953','72-3148','XX-XXXX']
FUEL_ROUTE = {'71-9538':2,'72-1620':2,'72-2237':1,'72-2420':1,'72-2953':1,'72-3148':1,'XX-XXXX':1}


def to_date(v):
    if v is None or v == '': return None
    if isinstance(v, datetime): return v.date().isoformat()
    if isinstance(v, date): return v.isoformat()
    return None


def to_num(v):
    if v is None or v == '': return None
    if isinstance(v, (int, float)): return float(v)
    try: return float(v)
    except (TypeError, ValueError): return None


def to_int(v):
    n = to_num(v)
    return int(n) if n is not None else None


def to_str(v):
    if v is None or v == '': return None
    return str(v).strip()


# ============================================================
# Loaders
# ============================================================

def load_outcomes():
    """Read OUT_COME (ใหญ่) — รายจ่ายทั้งฝูง"""
    wb = openpyxl.load_workbook(THIRA_XLSM, data_only=True, keep_vba=True, read_only=True)
    ws = wb['OUT_COME']
    rows = []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        dt, plate, cat, liters, amount, note = r[:6]
        if dt is None or plate is None or cat is None:
            continue
        rows.append({
            'date': to_date(dt),
            'truck_plate': to_str(plate),
            'category': to_str(cat),
            'liters': to_num(liters),
            'amount': to_num(amount),
            'note': to_str(note),
            'source_row': i,
        })
    wb.close()
    return rows


def load_incomes():
    """Read IN_COME — รายรับทั้งฝูง"""
    wb = openpyxl.load_workbook(THIRA_XLSM, data_only=True, keep_vba=True, read_only=True)
    ws = wb['IN_COME']
    rows = []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        dt, plate, gross, _net, note = r[:5]
        if dt is None or plate is None or gross is None:
            continue
        rows.append({
            'date': to_date(dt),
            'truck_plate': to_str(plate),
            'amount_gross': to_num(gross),
            'note': to_str(note),
            'source_row': i,
        })
    wb.close()
    return rows


def load_prices():
    """Read PRICE — ค่าเที่ยวคนขับ + ลิตรเป้า ต่อปลายทาง"""
    wb = openpyxl.load_workbook(THIRA_XLSM, data_only=True, keep_vba=True, read_only=True)
    ws = wb['PRICE']
    rows = []
    seen = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        dest = to_str(r[0])
        factory = to_str(r[1])
        if not dest:
            continue
        key = (dest, factory)
        if key in seen:
            continue
        seen.add(key)
        rows.append({
            'destination_code': dest,
            'destination_name': factory,
            'factory_origin': factory,
            'fee_driver': to_num(r[3]) or 0,
            'fuel_liters_route_1': to_num(r[4]),
            'fuel_liters_route_2': to_num(r[5]),
        })
    wb.close()
    return rows


PREFIX_TO_FACTORY = {
    'TWI':         'ทุ่งหลวง',
    'BSW':         'บ้านส้อง',
    'INV':         'ไทยนคร',
    'SHL':         'เอส.เอช.แอล (รัตภูมิ)',
    'พัทลุง':       'พัทลุงซัพพลาย',
    'ชะอวด':       'J.M. (ชะอวด)',
    'SJ-':         'สวนจันทร์ (ตรัง)',
    'กนกทอง':      'กนกทอง (เทพา)',
    'นาวาพารา':     'นาวาพารา',
    'นาเมืองเพชร':   'นาเมืองเพชร',
    'PTN':         'PTN40',
    'PTN20':       'PTN20',
    'APK':         'เอ.พี.เค.เฟอร์นิชชิ่ง',
    'วินวิน':       'วิน วิน',
}


def load_freight_price_sets():
    """Read PRICE_UPDATE → upsert ลง public.freight_price_sets (1 row ต่อ 1 วันที่)

    JSONB key = factory_name (mapped จาก prefix), value = ราคา
    id = 'ps_thira_YYYY-MM-DD'
    """
    wb = openpyxl.load_workbook(THIRA_XLSM, data_only=True, keep_vba=True, read_only=True)
    ws = wb['PRICE_UPDATE']
    header = [c.value for c in ws[1]]
    cols = []  # (col_idx, factory_name)
    for col_idx, h in enumerate(header):
        if h is None or col_idx == 0:
            continue
        prefix = to_str(h)
        factory = PREFIX_TO_FACTORY.get(prefix)
        if not factory:
            print(f'  ⚠️ unknown prefix in PRICE_UPDATE: {prefix!r} (col {col_idx}) — skipped')
            continue
        cols.append((col_idx, factory))
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        dt = r[0]
        if dt is None:
            continue
        d = to_date(dt)
        if not d:
            continue
        prices = {}
        for col_idx, factory in cols:
            price = to_num(r[col_idx])
            if price is None or price == 0:
                continue
            prices[factory] = price
        if not prices:
            continue
        rows.append({
            'id': f'ps_thira_{d}',
            'name': f'THIRA PRICE_UPDATE {d}',
            'prices': prices,
            'updated_at': datetime.combine(datetime.fromisoformat(d).date(), datetime.min.time()).isoformat(),
        })
    wb.close()
    return rows


def load_fuel_rates():
    """Pick up from INPUT!G2 (current rate) — single row seed"""
    wb = openpyxl.load_workbook(THIRA_XLSM, data_only=True, keep_vba=True, read_only=True)
    ws = wb['INPUT']
    rate = to_num(ws['G2'].value)
    wb.close()
    if rate is None:
        return []
    return [{'effective_date': date.today().isoformat(), 'baht_per_liter': rate}]


def load_containers():
    """Read DATABASE (ใบยกตู้V4) — งานวิ่งทุกคัน (ไม่ใช่แค่ THIRA)"""
    wb = openpyxl.load_workbook(ANTS_XLSM, data_only=True, keep_vba=True, read_only=True)
    ws = wb['DATABASE']
    rows = []
    for i, r in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        plate, bundles, inv, date_in, factory, vessel, cont_no, weight, cado, pickup, status = r[:11]
        if all(v in (None, '', 0) for v in (plate, inv, date_in, cont_no)):
            continue
        rows.append({
            'truck_plate': to_str(plate),
            'bundles': to_int(bundles),
            'invoice_no': to_str(inv),
            'date_in': to_date(date_in),
            'factory': to_str(factory),
            'vessel': to_str(vessel),
            'container_no': to_str(cont_no),
            'weight': to_num(weight),
            'cado_no': to_str(cado),
            'date_pickup': to_date(pickup),
            'status': to_str(status),
            'source_row': i,
        })
    wb.close()
    return rows


# ============================================================
# Pusher
# ============================================================

def chunked(seq, n):
    for i in range(0, len(seq), n):
        yield seq[i:i+n]


def push(supa, schema, table, rows, chunk_size=500, upsert_on=None):
    if not rows:
        print(f'  [{schema}.{table}] no rows')
        return 0
    total = 0
    for batch in chunked(rows, chunk_size):
        q = supa.schema(schema).table(table)
        if upsert_on:
            resp = q.upsert(batch, on_conflict=upsert_on).execute()
        else:
            resp = q.insert(batch).execute()
        total += len(resp.data) if resp.data else 0
    verb = 'upserted' if upsert_on else 'inserted'
    print(f'  [{schema}.{table}] {verb} {total} rows')
    return total


TABLES = {
    # name → (schema, table, loader, upsert_on)
    'prices':              ('thira',  'prices',              load_prices,              'destination_code,factory_origin'),
    'fuel_rates':          ('thira',  'fuel_rates',          load_fuel_rates,          'effective_date'),
    'incomes':             ('thira',  'incomes',             load_incomes,             None),
    'outcomes':            ('thira',  'outcomes',            load_outcomes,            None),
    'containers':          ('ants',   'containers',          load_containers,          None),
    # freight_price_sets: ใช้ของเดิมใน FREIGHT_CALC (ราคาตรง Excel แล้ว) — skip
    # 'freight_price_sets':  ('public', 'freight_price_sets',  load_freight_price_sets,  'id'),
}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry-run', action='store_true')
    ap.add_argument('--tables', default=','.join(TABLES.keys()))
    args = ap.parse_args()

    url = os.environ.get('SUPABASE_URL')
    key = os.environ.get('SUPABASE_SERVICE_KEY')
    if not args.dry_run and (not url or not key):
        print('ERROR: set SUPABASE_URL and SUPABASE_SERVICE_KEY env vars')
        sys.exit(1)

    supa: Client | None = None
    if not args.dry_run:
        supa = create_client(url, key)

    selected = [t.strip() for t in args.tables.split(',') if t.strip()]
    for name in selected:
        if name not in TABLES:
            print(f'  [skip] unknown table {name}')
            continue
        schema, table, loader, upsert_on = TABLES[name]
        print(f'Loading {schema}.{table}...')
        rows = loader()
        print(f'  loaded {len(rows)} rows from Excel')
        if rows:
            sample = {k: v for k, v in list(rows[0].items())[:6]}
            print(f'  sample: {sample}')
        if args.dry_run:
            continue
        push(supa, schema, table, rows, upsert_on=upsert_on)


if __name__ == '__main__':
    main()
